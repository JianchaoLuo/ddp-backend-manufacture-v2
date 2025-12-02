import redis
from openpyxl.reader.excel import load_workbook

import influxdb_client, os, time
from influxdb_client import InfluxDBClient, Point, WritePrecision
from influxdb_client.client.write_api import SYNCHRONOUS

token = "GQhOuQdCI6conK3sEZMugfQXQEGpUEYIRELAzoMjUnlku1fKYvMDvQupzLXTbQkJJFOW1p2G66q43te_LYJnOA=="

org = "hex"
url = "http://10.216.217.131:8086"

write_client = influxdb_client.InfluxDBClient(url=url, token=token, org=org)
bucket = "hex_bucket"

write_api = write_client.write_api(write_options=SYNCHRONOUS)

redis_client = redis.StrictRedis(host='10.216.217.131', username="default", password="KT4JzNNPEtXQmt", port=6379,
                                 db=1, )
mission_id = redis_client.get("MISSION_ID")
if not mission_id:
    # 如果设置，就设置为1
    redis_client.set("MISSION_ID", 1)
    mission_id = 1
else:
    mission_id = int(mission_id) + 1
    redis_client.incr("MISSION_ID", 1)
current_mission = mission_id
print("current_mission", current_mission)


class Equipment:
    def __init__(self, _id, status, current_action):
        self.id = _id
        self.status = status
        self.current_action = current_action

    def __str__(self):
        return f"PushEquipment({self.id}, EquipmentOperationStatusEnum.{self.status},\"{self.current_action}\"),"

    def toPoint(self, mission, tick):
        return (Point("equipment_real_time")
                .field("tick", tick)
                .tag("id", self.id)
                .tag("mission", mission)
                .field("operationStatus", self.status)
                .field("currentAction", self.current_action))


class Car:
    def __init__(self, _id, status, _type, coordinate, current_action):
        self.id = _id
        self.status = status
        self.type = _type
        self.coordinate = coordinate
        self.target = []
        # 如果是1，则x减，如果是2，则x加，如果是3，则y减，如果是4，则y加
        self.operation = 1
        self.current_action = current_action

    def __str__(self):
        if self.target:
            return f"PushCar({self.id}, Coordinate({self.coordinate[0]}, {self.coordinate[1]}), Coordinate({self.target[0]}, {self.target[1]}), \"{self.current_action}\"),"
        else:
            return f"PushCar({self.id}, Coordinate({self.coordinate[0]}, {self.coordinate[1]}), null, \"{self.current_action}\"),"

    def toPoint(self, mission, tick):
        cor_str = f"{self.coordinate[0]},{self.coordinate[1]}" if self.coordinate else ""
        target_str = f"{self.target[0]},{self.target[1]}" if self.target else ""
        return (Point("car_real_time")
                .field("tick", tick)
                .tag("id", self.id)
                .tag("mission", mission)
                .tag("type", self.type)
                .field("currentCoordinate", cor_str)
                .field("targetCoordinate", target_str)
                .field("operationStatus", self.status)
                .field("currentAction", self.current_action))


class Position:
    def __init__(self, _id, status, current_action):
        self.id = _id
        self.status = status
        self.current_action = current_action

    def __str__(self):
        return f"PositionRealTime({self.id}, PositionStatusEnum.{self.status}, \"{self.current_action}\"),"

    def toPoint(self, mission, tick):
        return (Point("position_real_time")
                .field("tick", tick)
                .tag("id", self.id)
                .tag("mission", mission)
                .field("status", self.status)
                .field("currentAction", self.current_action))


class Workstation:
    def __init__(self, _id, status, current_action):
        self.id = _id
        self.status = status
        self.current_action = current_action

    def __str__(self):
        return f"WorkstationRealTime({self.id}, WorkstationStatusEnum.{self.status}, \"{self.current_action}\"),"

    def toPoint(self, mission, tick):
        return (Point("workstation_real_time")
                .field("tick", tick)
                .tag("id", self.id)
                .tag("mission", mission)
                .field("status", self.status)
                .field("currentAction", self.current_action))


workbook = load_workbook(filename='仿真模拟数据-修正5.xlsx')
sheet = workbook.active
all_data = []
for i in range(4, 131):
    row_data = []
    for j in range(3, 124):
        row_data.append(sheet.cell(row=i, column=j).value)
    all_data.append(row_data)

equipments = [
    Equipment(1796862790497665025, "STANDBY", "待机"),
    Equipment(1796862790678020098, "STANDBY", "待机"),
    Equipment(1796862790879346689, "STANDBY", "待机"),
    Equipment(1796862790946455554, "STANDBY", "待机"),
    # 以下为新增
    Equipment(1796862790547996675, "STANDBY", "待机"),
    Equipment(1796862790745128961, "STANDBY", "待机"),
    Equipment(1796862791277805570, "STANDBY", "待机"),
    Equipment(1796862791344914435, "STANDBY", "待机"),
    Equipment(1796862790610911234, "STANDBY", "待机"),
    Equipment(1796862790812237825, "STANDBY", "待机")
]
workstations = [
    Workstation(1796862791412023297, "FREE", "空闲"),
    Workstation(1796862791474937858, "FREE", "空闲"),
    Workstation(1796862791537852419, "FREE", "空闲"),
    Workstation(1796862791604961283, "FREE", "空闲"),
    # 以下为新增
    Workstation(1796862791672070147, "FREE", "空闲"),
    Workstation(1796862791760150529, "FREE", "空闲"),
    Workstation(1796862791802093571, "FREE", "空闲"),
    Workstation(1796862791869202433, "FREE", "空闲")
]
positions = [
    Position(1796864626461900801, "UNOCCUPIED", "未占用"),
    Position(1796864626629672962, "UNOCCUPIED", "未占用"),
    # 以下为新增
    Position(1796864626499649538, "UNOCCUPIED", "未占用"),
    Position(1796864626696781826, "UNOCCUPIED", "未占用"),
    Position(1806222687347474434, "UNOCCUPIED", "未占用"),
    Position(1796864626763890689, "UNOCCUPIED", "未占用"),
    Position(1796864626826805250, "UNOCCUPIED", "未占用"),
    Position(1796864626961022977, "UNOCCUPIED", "未占用"),
    Position(1796864626893914113, "UNOCCUPIED", "未占用"),
    Position(1796864626961022979, "UNOCCUPIED", "未占用"),
    Position(1806222939316092929, "UNOCCUPIED", "未占用"),
    Position(1796864627296567297, "UNOCCUPIED", "未占用"),
    Position(1796864627422396417, "UNOCCUPIED", "未占用"),
    Position(1806767877643046913, "UNOCCUPIED", "未占用"),
    Position(1806767969666076674, "UNOCCUPIED", "未占用")
]
cars = [
    Car(1, "WORKING", "FERRY_CAR", [3410, 1404], "运载子车(id:914435)"),
    Car(2, "WORKING", "FERRY_CAR", [3238, 955], "运载子车(id:914965)"),
    Car(3, "WORKING", "FERRY_CAR", [2048, 969], "无运载"),
    Car(4, "WORKING", "SUB_CAR", [3410, 1404], "无运载"),
    Car(5, "WORKING", "SUB_CAR", [3238, 955], "无运载"),
    Car(6, "WORKING", "SUB_CAR", [2929, 865], "无运载"),
    # 冷砂合模轨道母车
    Car(1806898212104048642, "WORKING", "FERRY_CAR", [3110, 1194], "无运载"),
    # 冷砂合模轨道子车
    Car(1806898464978636801, "WORKING", "SUB_CAR", [3110, 1194], "无运载"),
    # 制芯轨道子车
    Car(1806222141957931009, "WORKING", "SUB_CAR", [3321, 1013], "无运载"),
    # 模具回流轨道子车
    Car(1806222498566045698, "WORKING", "SUB_CAR", [3311, 1132], "无运载"),
    # 热砂合箱垂直轨道字车
    Car(1806898664916914177, "WORKING", "SUB_CAR", [2048, 969], "无运载"),
    # 冷砂射砂轨道子车
    Car(1806898891212197890, "WORKING", "SUB_CAR", [2583, 1248], "无运载"),
    # 冷砂合箱垂直轨道母车
    Car(1806899040621694978, "WORKING", "FERRY_CAR", [2048, 1334], "无运载"),
    # 冷砂合箱垂直轨道子车
    Car(1806899108238069761, "WORKING", "SUB_CAR", [2048, 1334], "无运载"),
    # 模具回流垂直轨道子车
    Car(1806899247602208770, "WORKING", "SUB_CAR", [2443, 992], "无运载"),
    # 冷砂模具回流垂直轨道子车
    Car(1806899447880224769, "WORKING", "SUB_CAR", [2380, 1192], "无运载"),
    # 冷砂砂箱喷涂垂直轨道子车
    Car(1806899510102724610, "WORKING", "SUB_CAR", [2306, 1312], "无运载"),
    # 冷砂砂箱喷涂轨道子车
    Car(1806899642168774658, "WORKING", "SUB_CAR", [2256, 1372], "无运载")
]
now_time = 0
_index = 0
speed = 30
all_tick = 0
while True:
    try:
        row = all_data[_index]
    except:
        break
    _index += 1
    time = row[0]
    if time < 0.5:
        continue
    ticks = int(time / 0.5)
    # 如果是1，则x减，如果是2，则x加，如果是3，则y减，如果是4，则y加
    if row[67]:
        row11 = row[67].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        cars[0].target = [int(row11[0]), int(row11[1])]
        if row11[0] == cars[0].coordinate[0]:
            # x相同
            cars[0].operation = 3 if int(row11[1]) < cars[0].coordinate[1] else 4
        else:
            # y相同
            cars[0].operation = 1 if int(row11[0]) < cars[0].coordinate[0] else 2
    if row[70]:
        row11 = row[70].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        cars[3].target = [int(row11[0]), int(row11[1])]
        if row11[0] == cars[3].coordinate[0]:
            # x相同
            cars[3].operation = 3 if int(row11[1]) < cars[3].coordinate[1] else 4
        else:
            # y相同
            cars[3].operation = 1 if int(row11[0]) < cars[3].coordinate[0] else 2
    if row[73]:
        row11 = row[73].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[1]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    if row[76]:
        row11 = row[76].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[4]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    if row[79]:
        row11 = row[79].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[5]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    if row[82]:
        row11 = row[82].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[2]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    # 以下新加
    if row[85]:
        row11 = row[85].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[6]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    if row[88]:
        row11 = row[88].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[7]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    if row[91]:
        row11 = row[91].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[8]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    if row[94]:
        row11 = row[94].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[9]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    if row[97]:
        row11 = row[97].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[10]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    if row[100]:
        row11 = row[100].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[11]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    if row[103]:
        row11 = row[103].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[12]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    if row[106]:
        row11 = row[106].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[13]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    if row[109]:
        row11 = row[109].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[14]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    if row[112]:
        row11 = row[112].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[15]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    if row[115]:
        row11 = row[115].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[16]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    if row[118]:
        row11 = row[118].strip().split(", ")
        row11 = [int(row11[0]), int(row11[1])]
        car = cars[17]
        car.target = [int(row11[0]), int(row11[1])]
        if row11[0] == car.coordinate[0]:
            # x相同
            car.operation = 3 if int(row11[1]) < car.coordinate[1] else 4
        else:
            # y相同
            car.operation = 1 if int(row11[0]) < car.coordinate[0] else 2
    for i in range(ticks):
        for car in cars:
            if car.target:
                # 如果是1，则x减，如果是2，则x加，如果是3，则y减，如果是4，则y加
                if car.operation == 1 and car.coordinate[0] > car.target[0]:
                    car.coordinate[0] -= speed
                elif car.operation == 2 and car.coordinate[0] < car.target[0]:
                    car.coordinate[0] += speed
                elif car.operation == 3 and car.coordinate[1] > car.target[1]:
                    car.coordinate[1] -= speed
                elif car.operation == 4 and car.coordinate[1] < car.target[1]:
                    car.coordinate[1] += speed

        if i == ticks - 1:
            for car in cars:
                if car.target:
                    car.coordinate = car.target
                    car.target = []

            if row[1]:
                equipments[0].status = row[1]
            if row[2]:
                equipments[0].current_action = row[2]
            if row[3]:
                equipments[1].status = row[3]
            if row[4]:
                equipments[1].current_action = row[4]
            if row[5]:
                equipments[2].status = row[5]
            if row[6]:
                equipments[2].current_action = row[6]
            if row[7]:
                equipments[3].status = row[7]
            if row[8]:
                equipments[3].current_action = row[8]
            # 新加的设备
            for row_index in range(9, 21, 2):
                if row[row_index]:
                    equipments[row_index // 2].status = row[row_index]
                if row[row_index + 1]:
                    equipments[row_index // 2].current_action = row[row_index + 1]
            # 所有工岗
            entity_index = 0
            for row_index in range(21, 37, 2):
                if row[row_index]:
                    workstations[entity_index].status = row[row_index]
                if row[row_index + 1]:
                    workstations[entity_index].current_action = row[row_index + 1]
                entity_index += 1
            # 所有点位
            entity_index = 0
            for row_index in range(37, 67, 2):
                if row[row_index]:
                    positions[entity_index].status = row[row_index]
                if row[row_index + 1]:
                    positions[entity_index].current_action = row[row_index + 1]
                entity_index += 1
            # if row[9]:
            #     workstations[0].status = row[9]
            # if row[10]:
            #     workstations[0].current_action = row[10]
            # if row[11]:
            #     workstations[1].status = row[11]
            # if row[12]:
            #     workstations[1].current_action = row[12]
            # if row[13]:
            #     workstations[2].status = row[13]
            # if row[14]:
            #     workstations[2].current_action = row[14]
            # if row[15]:
            #     workstations[3].status = row[15]
            # if row[16]:
            #     workstations[3].current_action = row[16]

            # if row[17]:
            #     positions[0].status = row[17]
            # if row[18]:
            #     positions[0].current_action = row[18]
            # if row[19]:
            #     positions[1].status = row[19]
            # if row[20]:
            #     positions[1].current_action = row[20]

            if row[67]:
                cars[0].current_action = row[67]
            if row[70]:
                cars[3].current_action = row[70]
            if row[73]:
                cars[1].current_action = row[73]
            if row[76]:
                cars[4].current_action = row[76]
            if row[79]:
                cars[5].current_action = row[79]
            if row[82]:
                cars[2].current_action = row[82]
            # 新加的车辆
            entity_index = 6
            for row_index in range(85, 121, 3):
                if row[row_index]:
                    cars[entity_index].current_action = row[row_index]
                entity_index += 1

        # print(
        #     f"listOf<Any>(listOf({''.join([str(i) for i in equipments])}),listOf({''.join([str(i) for i in workstations])}),listOf({''.join([str(i) for i in positions])}),listOf({''.join([str(i) for i in cars])}),),"
        # )
        for j in cars:
            write_api.write(bucket=bucket, org="hex", record=j.toPoint(current_mission, all_tick))
        for j in equipments:
            write_api.write(bucket=bucket, org="hex", record=j.toPoint(current_mission, all_tick))
        for j in workstations:
            write_api.write(bucket=bucket, org="hex", record=j.toPoint(current_mission, all_tick))
        for j in positions:
            write_api.write(bucket=bucket, org="hex", record=j.toPoint(current_mission, all_tick))

        all_tick += 1
redis_client.set("MISSION_MAX_TICK", all_tick - 1)
print("MISSION_MAX_TICK", all_tick - 1)
